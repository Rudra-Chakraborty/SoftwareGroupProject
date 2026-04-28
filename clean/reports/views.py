from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import Department, Team, TeamMember



def login_required_session(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_email'):
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


def _get_staff(request):
    try:
        from main.models import User, Staff
        email = request.session.get('user_email')
        user = User.objects.get(email=email)
        return Staff.objects.filter(user=user).first()
    except:
        return None


def _get_report_data():
    departments = Department.objects.all()
    all_teams = Team.objects.select_related('department').prefetch_related('members')
    inactive_teams = all_teams.filter(is_active=False)

    dept_summary = []
    for dept in departments:
        teams = all_teams.filter(department=dept)
        dept_summary.append({
            'name': dept.name or 'Unknown',
            'team_count': teams.count(),
            'head': dept.leader or 'N/A',
        })

    return {
        'departments': departments,
        'all_teams': all_teams,
        'teams_without_managers': inactive_teams,
        'total_teams': all_teams.count(),
        'total_departments': departments.count(),
        'total_members': TeamMember.objects.count(),
        'dept_summary': dept_summary,
    }


@login_required_session
def reports_dashboard(request):
    data = _get_report_data()
    data['staff'] = _get_staff(request)
    return render(request, 'reports/dashboard.html', data)


@login_required_session
def generate_pdf(request):
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("ReportLab not installed.", status=500)

    data = _get_report_data()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="skytrack_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Title'], fontSize=22, textColor=colors.HexColor('#1a56db'), spaceAfter=6, alignment=TA_CENTER)
    h2 = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1e429f'), spaceBefore=14, spaceAfter=6)

    story = []
    story.append(Paragraph("Sky Track – Engineering Teams Report", title_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a56db')))
    story.append(Spacer(1, 0.3*inch))

    story.append(Paragraph("Summary", h2))
    t = Table([['Metric','Value'],['Total Departments',str(data['total_departments'])],['Total Teams',str(data['total_teams'])],['Total Members',str(data['total_members'])],['Inactive Teams',str(data['teams_without_managers'].count())]], colWidths=[10*cm,6*cm])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a56db')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),10),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor('#f0f4ff'),colors.white]),('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#d1d5db')),('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),8)]))
    story.append(t)
    story.append(Spacer(1, 0.3*inch))

    story.append(Paragraph("All Teams", h2))
    rows = [['Team','Department','Manager','Members','Active']] + [[t.name, t.department.name or 'N/A', t.manager, str(t.members.count()), 'Yes' if t.is_active else 'No'] for t in data['all_teams']]
    tt = Table(rows, colWidths=[4*cm,4*cm,4*cm,3*cm,2*cm])
    tt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e429f')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor('#f0f4ff'),colors.white]),('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#d1d5db')),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),('LEFTPADDING',(0,0),(-1,-1),8)]))
    story.append(tt)
    doc.build(story)
    return response


@login_required_session
def generate_excel(request):
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl not installed.", status=500)

    data = _get_report_data()
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color='FFFFFF', size=11)
    hb = PatternFill('solid', fgColor='1A56DB')
    hg = PatternFill('solid', fgColor='057A55')
    ca = Alignment(horizontal='center', vertical='center')
    la = Alignment(horizontal='left', vertical='center')
    thin = Side(border_style='thin', color='D1D5DB')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    fa = PatternFill('solid', fgColor='EBF5FF')
    fb = PatternFill('solid', fgColor='FFFFFF')

    def sh(ws, row, cols, fill):
        for col in range(1, cols+1):
            c = ws.cell(row=row, column=col); c.font=hf; c.fill=fill; c.alignment=ca; c.border=bd

    def sr(ws, row, cols, idx):
        fill = fa if idx%2==0 else fb
        for col in range(1, cols+1):
            c = ws.cell(row=row, column=col); c.fill=fill; c.border=bd; c.alignment=la

    ws1 = wb.active; ws1.title = "Summary"
    ws1['A1'] = 'Sky Track Report'; ws1['A1'].font = Font(bold=True, size=16, color='1A56DB')
    ws1.merge_cells('A1:B1'); ws1['A1'].alignment = ca
    ws1.append([]); ws1.append(['Metric','Value']); sh(ws1,3,2,hb)
    for i,(m,v) in enumerate([('Total Departments',data['total_departments']),('Total Teams',data['total_teams']),('Total Members',data['total_members']),('Inactive Teams',data['teams_without_managers'].count())]):
        ws1.append([m,v]); sr(ws1,4+i,2,i)
    ws1.column_dimensions['A'].width=30; ws1.column_dimensions['B'].width=15

    ws2 = wb.create_sheet("All Teams")
    h2 = ['Team Name','Department','Manager','Email','Members','Active']; ws2.append(h2); sh(ws2,1,len(h2),hb)
    for i,t in enumerate(data['all_teams']):
        ws2.append([t.name,t.department.name or 'N/A',t.manager,t.contact_email,t.members.count(),'Yes' if t.is_active else 'No']); sr(ws2,i+2,len(h2),i)
    for ci,w in zip(range(1,7),[25,20,20,25,12,10]): ws2.column_dimensions[get_column_letter(ci)].width=w

    ws3 = wb.create_sheet("Departments")
    h3 = ['Department','Leader','Specialisation','Teams']; ws3.append(h3); sh(ws3,1,len(h3),hg)
    for i,d in enumerate(data['dept_summary']):
        try: dept_obj = Department.objects.get(name=d['name']); spec = dept_obj.specialisation
        except: spec = ''
        ws3.append([d['name'],d['head'],spec,d['team_count']]); sr(ws3,i+2,len(h3),i)
    for ci,w in zip(range(1,5),[25,20,25,15]): ws3.column_dimensions[get_column_letter(ci)].width=w

    if data['dept_summary']:
        chart = BarChart(); chart.type='col'; chart.title='Teams per Department'; chart.style=10; chart.width=18; chart.height=12
        dr = Reference(ws3,min_col=4,min_row=1,max_row=len(data['dept_summary'])+1)
        cats = Reference(ws3,min_col=1,min_row=2,max_row=len(data['dept_summary'])+1)
        chart.add_data(dr,titles_from_data=True); chart.set_categories(cats); ws3.add_chart(chart,"F2")

    ws4 = wb.create_sheet("All Members")
    h4 = ['Name','Role','Email','Team','Department']; ws4.append(h4); sh(ws4,1,len(h4),hb)
    for i,m in enumerate(TeamMember.objects.select_related('team__department').all()):
        ws4.append([m.name,m.role,m.email,m.team.name,m.team.department.name]); sr(ws4,i+2,len(h4),i)
    for ci,w in zip(range(1,6),[25,20,25,20,20]): ws4.column_dimensions[get_column_letter(ci)].width=w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="skytrack_report.xlsx"'
    wb.save(response)
    return response


@login_required_session
def chart_data(request):
    data = _get_report_data()
    return JsonResponse({
        'labels': [d['name'] for d in data['dept_summary']],
        'team_counts': [d['team_count'] for d in data['dept_summary']],
        'with_managers': [d['team_count'] for d in data['dept_summary']],
        'without_managers': [0 for d in data['dept_summary']],
        'heads': [d['head'] for d in data['dept_summary']],
        'total_teams': data['total_teams'],
        'total_departments': data['total_departments'],
        'total_members': data['total_members'],
        'no_manager_count': data['teams_without_managers'].count(),
    })
